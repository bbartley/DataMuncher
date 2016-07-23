import openpyxl as xlsx
import numpy as np
import numpy.ma as mask

class Experiment:
    def __init__(self, xlsx_file, protocol=None):
        self.workbook = xlsx.load_workbook(xlsx_file)
        self.load_protocol(protocol)

    # The protocol specifies a callback function for parsing a plate reader file in different formats
    def load_protocol(self, protocol):
        if protocol:
            protocol(self)

    # Search for excel cell value
    def search(self, cell_value):
        ws = self.workbook.active
        for row in range(1, ws.max_row):
            for col in range(1, ws.max_column):
                if ws.cell(row=row, column=col).value == cell_value:
                    return ws.cell(row=row, column=col)
        return None

    # Read a column of data given the column header
    def read(self, header_label, direction=1):  # 0 for horizontal, 1 for vertical
        ws = self.workbook.active
        header_cell = self.search(header_label)
        j = xlsx.utils.column_index_from_string(header_cell.column)
        current_cell = ws.cell(row=header_cell.row + 1, column=j)
        scanned_cells = []
        while not current_cell.value == None:
            scanned_cells.append(current_cell)
            current_cell = ws.cell(row=current_cell.row + 1, column=j)
        return scanned_cells

class DataSeries():

    def __init__(self, data_series = None):
        if not data_series:
            self.list_of_series = []
        else:
            self.list_of_series = [ data_series ]

    def __len__(self):
        return len(self.list_of_series)

    def append(self, data_series):
        self.list_of_series.extend(data_series.list_of_series)
        #for s in data_series.list_of_series:
        #    self.list_of_series.append(s)
        print(self.list_of_series)

class MeasurementGroup():
    def __init__(self):
        self.object_map = {}  # Should be of the form { '' : MeasurementGroup}
        self.cell_map = {}  # Should be of the form { '' : MaskedArray([xlsx cells])}
        #self.measurements = []  # A list of list of xlsx_cells
        self.measurements = DataSeries()

    # def get_measurements(self, data_series = None):
    #     if self.measurements:
    #         if not data_series:
    #             data_series = [ self.measurements ]
    #         else:
    #             data_series.append(self.measurements)
    #         return data_series
    #     else:
    #         for obj in self.object_map.values():
    #             data_series = obj.get_measurements(data_series)
    #         return data_series

    def get_measurements(self, data_series = None):
        if len(self.measurements):
            if not data_series:
                data_series = self.measurements
            else:
                data_series.append(self.measurements)
            return data_series
        else:
            for obj in self.object_map.values():
                data_series = obj.get_measurements(data_series)
            return data_series

    def __getitem__(self, key):
        if isinstance(key, slice):
            raise Exception('Slice indices not implemented')
        elif isinstance(key, str):
            return self.object_map[str]
        else: # assume int-like object
            if key < 0: # if negative index, convert to positive and start from end
                key += len(self)
            raise Exception('Numerical indices not implemented')

class Groups(MeasurementGroup):
    def __init__(self):
        pass

class Samples(MeasurementGroup):
    def __init__(self):
        MeasurementGroup.__init__(self)

class Well(MeasurementGroup):
    def __init__(self):
        MeasurementGroup.__init__(self)

f = '160514 growth curves.xlsx'
ex = Experiment(f)
list_of_xlsx_cells = ex.read('Time [s]')
time_labels = [cel.value for cel in list_of_xlsx_cells]
S = Samples()

row_ids = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
column_ids = map(str, range(1, 13))
well_ids = [ r + c for r in row_ids for c in column_ids if ex.search(r + c)]  # Get well IDs from data file
for w in well_ids:
    W = Well()
    W.measurements = DataSeries(ex.read(w))  # Read well data vertically in columns from plate reader spreadsheet
    S.object_map[w] = W
data_series = S.get_measurements()

for s in data_series.list_of_series:
    print s
    raw_input()