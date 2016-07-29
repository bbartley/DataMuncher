import openpyxl as xlsx
import numpy as np
import numpy.ma as mask
import matplotlib.pyplot as plt

class Experiment:
    def __init__(self, xlsx_file, protocol=None):
        self.workbook = xlsx.load_workbook(xlsx_file)
        self.load_protocol(protocol)

    # The protocol specifies a callback function for parsing a plate reader file in different formats
    def load_protocol(self, protocol):
        if protocol:
            protocol(self)

    # Search for excel cell value
    def search(self, cell_value):
        ws = self.workbook.active
        for row in range(1, ws.max_row):
            for col in range(1, ws.max_column):
                if ws.cell(row=row, column=col).value == cell_value:
                    return ws.cell(row=row, column=col)
        return None

    # Read a column of data given the column header
    def read(self, header_label, direction=1):  # 0 for horizontal, 1 for vertical
        ws = self.workbook.active
        header_cell = self.search(header_label)
        j = xlsx.utils.column_index_from_string(header_cell.column)
        current_cell = ws.cell(row=header_cell.row + 1, column=j)
        scanned_cells = []
        while not current_cell.value == None:
            scanned_cells.append(current_cell.value)
            current_cell = ws.cell(row=current_cell.row + 1, column=j)
        return DataSeries(scanned_cells)

class DataSeries():

    def __init__(self, data_series = None):
        if not data_series:
            self.list_of_series = []
        else:
            self.list_of_series = [ data_series ]

    def __len__(self):
        return len(self.list_of_series)

    def append(self, data_series):
        self.list_of_series.extend(data_series.list_of_series)
        #for s in data_series.list_of_series:
        #    self.list_of_series.append(s)

    def group_measurements(self):
        zipped_data_series = []
        for i_measurement in range(0, len(self.list_of_series[0])):
            grouped_measurements = []
            for i_series in range(0, len(self.list_of_series)):
                data_point = self.list_of_series[i_series][i_measurement]
                grouped_measurements.append(data_point)
            zipped_data_series.append(grouped_measurements)
        return zipped_data_series

    def mean(self):
        groups = self.group_measurements()
        mean_measurements = [sum(g) / len(g) for g in groups]
        return mean_measurements


    def std(self):
        groups = self.group_measurements()
        stdevs = [ np.std(g) for g in groups]
        return stdevs

class MeasurementGroup():
    def __init__(self):
        self.object_map = {}  # Should be of the form { '' : MeasurementGroup}
        self.cell_map = {}  # Should be of the form { '' : MaskedArray([xlsx cells])}
        #self.measurements = []  # A list of list of xlsx_cells
        self.measurements = DataSeries()

    # def get_measurements(self, data_series = None):
    #     if self.measurements:
    #         if not data_series:
    #             data_series = [ self.measurements ]
    #         else:
    #             data_series.append(self.measurements)
    #         return data_series
    #     else:
    #         for obj in self.object_map.values():
    #             data_series = obj.get_measurements(data_series)
    #         return data_series

    def get_measurements(self, data_series = None):
        if len(self.measurements):
            if not data_series:
                data_series = self.measurements
            else:
                data_series.append(self.measurements)
            return data_series
        else:
            for obj in self.object_map.values():
                data_series = obj.get_measurements(data_series)
            return data_series

    def __getitem__(self, key):
        if isinstance(key, slice):
            raise Exception('Slice indices not implemented')
        elif isinstance(key, str):
            return self.object_map[key]
        else: # assume int-like object
            if key < 0: # if negative index, convert to positive and start from end
                key += len(self)
            raise Exception('Numerical indices not implemented')

    def __repr__(self):
        # TODO: if object_map points to DataSeries then return DataSeries.list_of_series
        return str(self.object_map.keys())

    def __str__(self):
        # TODO: if object_map points to DataSeries then return DataSeries.list_of_series
        return str(self.object_map.keys())

    def mean(self):
        return self.get_measurements().mean()

    def std(self):
        return self.get_measurements().std()

class Group(MeasurementGroup):
    def __init__(self):
        MeasurementGroup.__init__(self)

    def addSample(self, sample_name):
        self.object_map[sample_name] = Sample()

class Sample(MeasurementGroup):
    def __init__(self, list_of_well_ids = None):
        MeasurementGroup.__init__(self)
        if list_of_well_ids:
            for w in list_of_well_ids:
                self.object_map[w] = Well()

    def addWell(self, well_id, well_obj):
        self.object_map[well_id] = well_obj

class Well(MeasurementGroup):
    def __init__(self):
        MeasurementGroup.__init__(self)

# Load data file from plate reader
f = '160514 growth curves.xlsx'
ex = Experiment(f)
data_series = ex.read('Time [s]')
time_labels = data_series.list_of_series[0]

# Search plate reader file by well ID and associate a DataSeries with each Well
PlateReadout = Sample()
row_ids = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
column_ids = map(str, range(1, 13))
well_ids = [ r + c for r in row_ids for c in column_ids if ex.search(r + c)]  # Get well IDs from data file
for w in well_ids:
    W = Well()
    W.measurements = ex.read(w)  # Read well data vertically in columns from plate reader spreadsheet
    PlateReadout.addWell(w, W)
data_series = PlateReadout.get_measurements()

# List all the Wells in the data file
print(PlateReadout)


# Assign Samples from plate layout
Culture1 = Sample()
Culture1.addWell('B10', PlateReadout['B10'])
Culture1.addWell('B11', PlateReadout['B11'])
Culture1.addWell('B12', PlateReadout['B12'])

# Print Sample statistics
print( Culture1.mean() )
print( Culture1.std() )

# Graph the growth curve
#plt.plot(time_labels, Culture1.mean())
#plt.show()

# TODO
# change list operations to MaskedArray operations
# automate reading of plate layout
# change the return value of Experiment.read() to a DataSeries object
# change to object_map to an OrderedDict to support representation of dose levels
